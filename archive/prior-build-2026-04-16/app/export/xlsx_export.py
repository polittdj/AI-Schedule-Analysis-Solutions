"""Excel data export.

Generates a multi-sheet `.xlsx` workbook containing every piece of
engine output in tabular form. The workbook is meant for analysts who
want to sort, filter, and pivot the data themselves rather than read
the narrative report — it is the raw numeric backbone of the
forensic analysis.

Sheet layout
------------
* **Summary** — project metadata + high-level engine metrics
* **All Tasks** — full task list
* **Critical Path** — critical-path tasks only
* **Slippage** — comparator deltas (only for comparative runs)
* **DCMA Scorecard** — 14 metrics with conditional PASS / FAIL coloring
* **Manipulation** — every finding with confidence + category

Every sheet has a styled header row, auto-sized columns, and frozen
panes on the header row.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1B2432", end_color="1B2432", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)

PASS_FILL = PatternFill(start_color="D4F5E3", end_color="D4F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FBD7DB", end_color="FBD7DB", fill_type="solid")
PASS_FONT = Font(name="Arial", bold=True, color="1A7A46")
FAIL_FONT = Font(name="Arial", bold=True, color="9C1E2C")

HIGH_FILL = PatternFill(start_color="FBD7DB", end_color="FBD7DB", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FCEBC8", end_color="FCEBC8", fill_type="solid")
LOW_FILL = PatternFill(start_color="D6E9FC", end_color="D6E9FC", fill_type="solid")

MAX_COL_WIDTH = 50


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #


def generate_xlsx_export(results: Dict[str, Any]) -> bytes:
    """Build the full workbook and return it as bytes."""
    wb = Workbook()
    _write_summary(wb.active, results)
    wb.active.title = "Summary"

    schedule = _primary_schedule(results)
    if schedule is not None:
        _write_all_tasks(wb.create_sheet("All Tasks"), schedule)
        if results.get("cpm") is not None:
            _write_critical_path(
                wb.create_sheet("Critical Path"), results["cpm"], schedule
            )

    if results.get("comparison") is not None:
        _write_slippage(wb.create_sheet("Slippage"), results["comparison"])

    if results.get("dcma") is not None:
        _write_dcma(wb.create_sheet("DCMA Scorecard"), results["dcma"])

    if results.get("manipulation") is not None:
        _write_manipulation(
            wb.create_sheet("Manipulation"), results["manipulation"]
        )

    if results.get("trend") is not None:
        _write_trend(wb.create_sheet("Trend"), results["trend"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


def _primary_schedule(results: Dict[str, Any]):
    return results.get("later_schedule") or results.get("prior_schedule")


def _fmt_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(value)


def _style_header_row(ws: Worksheet, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 22
    # Assign the coordinate string directly; accessing ws["A2"] before
    # row 2 exists would materialize a blank cell and add a phantom row.
    ws.freeze_panes = f"A{row_num + 1}"


def _autosize_columns(ws: Worksheet) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        width = min(max_length + 2, MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 8)


# --------------------------------------------------------------------------- #
# Sheets
# --------------------------------------------------------------------------- #


def _write_summary(ws: Worksheet, results: Dict[str, Any]) -> None:
    ws.append(["Field", "Value"])
    _style_header_row(ws)

    schedule = _primary_schedule(results)
    info = schedule.project_info if schedule is not None else None
    has_comparison = results.get("comparison") is not None

    rows: List[List[Any]] = [
        [
            "Report type",
            "Comparative analysis" if has_comparison else "Single schedule",
        ],
        ["Generated (UTC)", datetime.utcnow().isoformat(timespec="seconds")],
    ]

    if info is not None:
        rows.extend(
            [
                ["Project name", info.name or "—"],
                ["Start date", _fmt_date(info.start_date) or "—"],
                ["Finish date", _fmt_date(info.finish_date) or "—"],
                ["Status date", _fmt_date(info.status_date) or "—"],
                ["Task count", len(schedule.tasks)],
            ]
        )

    cpm = results.get("cpm")
    if cpm is not None:
        rows.append(["Project duration (d)", round(cpm.project_duration_days, 2)])
        rows.append(["Critical path length", len(cpm.critical_path_uids)])

    dcma = results.get("dcma")
    if dcma is not None:
        rows.append(["DCMA passed", dcma.passed_count])
        rows.append(["DCMA failed", dcma.failed_count])
        rows.append(["DCMA overall %", round(dcma.overall_score_pct, 1)])

    comparison = results.get("comparison")
    if comparison is not None:
        rows.extend(
            [
                ["Tasks slipped", comparison.tasks_slipped_count],
                ["Tasks pulled in", comparison.tasks_pulled_in_count],
                ["Tasks completed", comparison.tasks_completed_count],
                ["Tasks added", comparison.tasks_added_count],
                ["Tasks deleted", comparison.tasks_deleted_count],
                ["Baseline movements", comparison.baseline_movement_count],
                [
                    "Completion slip (cal d)",
                    comparison.completion_date_slip_days,
                ],
            ]
        )

    manip = results.get("manipulation")
    if manip is not None:
        rows.extend(
            [
                ["Manipulation score", round(manip.overall_score, 1)],
                ["HIGH findings", manip.confidence_summary.get("HIGH", 0)],
                ["MEDIUM findings", manip.confidence_summary.get("MEDIUM", 0)],
                ["LOW findings", manip.confidence_summary.get("LOW", 0)],
            ]
        )

    ev = results.get("earned_value")
    if ev is not None:
        rows.extend(
            [
                ["EV units", ev.units],
                ["BAC", round(ev.budget_at_completion, 2)],
                ["PV", round(ev.planned_value, 2)],
                ["EV", round(ev.earned_value, 2)],
                ["SV", round(ev.schedule_variance, 2)],
                ["SPI", round(ev.schedule_performance_index, 4)],
            ]
        )

    fa = results.get("float_analysis")
    if fa is not None:
        rows.extend(
            [
                ["Float trend", fa.trend],
                ["Net float delta (d)", round(fa.net_float_delta, 2)],
                ["Became critical count", len(fa.became_critical_uids)],
                [
                    "Dropped off critical count",
                    len(fa.dropped_off_critical_uids),
                ],
            ]
        )

    for row in rows:
        ws.append(row)

    _autosize_columns(ws)


def _write_all_tasks(ws: Worksheet, schedule) -> None:
    ws.append(
        [
            "UID",
            "ID",
            "Name",
            "WBS",
            "Outline Level",
            "Start",
            "Finish",
            "Duration (d)",
            "Baseline Start",
            "Baseline Finish",
            "Baseline Duration (d)",
            "Actual Start",
            "Actual Finish",
            "% Complete",
            "Remaining Duration (d)",
            "Total Slack (d)",
            "Free Slack (d)",
            "Critical",
            "Summary",
            "Milestone",
            "Constraint Type",
            "Resources",
            "Predecessors",
            "Successors",
        ]
    )
    _style_header_row(ws)

    for t in schedule.tasks:
        ws.append(
            [
                t.uid,
                t.id,
                t.name,
                t.wbs,
                t.outline_level,
                _fmt_date(t.start),
                _fmt_date(t.finish),
                t.duration,
                _fmt_date(t.baseline_start),
                _fmt_date(t.baseline_finish),
                t.baseline_duration,
                _fmt_date(t.actual_start),
                _fmt_date(t.actual_finish),
                t.percent_complete,
                t.remaining_duration,
                t.total_slack,
                t.free_slack,
                "Y" if t.critical else "",
                "Y" if t.summary else "",
                "Y" if t.milestone else "",
                t.constraint_type,
                t.resource_names,
                ", ".join(str(p) for p in t.predecessors),
                ", ".join(str(s) for s in t.successors),
            ]
        )

    _autosize_columns(ws)


def _write_critical_path(ws: Worksheet, cpm, schedule) -> None:
    ws.append(
        [
            "Sequence",
            "UID",
            "Task",
            "Start",
            "Finish",
            "Duration (d)",
            "% Complete",
            "Early Start",
            "Early Finish",
            "Late Start",
            "Late Finish",
            "Total Float",
            "Free Float",
        ]
    )
    _style_header_row(ws)

    task_by_uid = {t.uid: t for t in schedule.tasks}
    float_lookup = cpm.task_floats or {}
    for idx, uid in enumerate(cpm.critical_path_uids, start=1):
        task = task_by_uid.get(uid)
        tf = float_lookup.get(uid)
        if task is None:
            continue
        ws.append(
            [
                idx,
                uid,
                task.name,
                _fmt_date(task.start),
                _fmt_date(task.finish),
                task.duration,
                task.percent_complete,
                getattr(tf, "early_start", None),
                getattr(tf, "early_finish", None),
                getattr(tf, "late_start", None),
                getattr(tf, "late_finish", None),
                getattr(tf, "total_float", None),
                getattr(tf, "free_float", None),
            ]
        )

    _autosize_columns(ws)


def _write_slippage(ws: Worksheet, comparison) -> None:
    ws.append(
        [
            "UID",
            "Task",
            "Start Slip (d)",
            "Finish Slip (d)",
            "Duration Change (d)",
            "% Complete Δ",
            "Total Slack Δ",
            "Baseline Start Δ (d)",
            "Baseline Finish Δ (d)",
            "Became Critical",
            "Dropped off Critical",
        ]
    )
    _style_header_row(ws)

    for d in comparison.task_deltas:
        ws.append(
            [
                d.uid,
                d.name,
                d.start_slip_days,
                d.finish_slip_days,
                d.duration_change_days,
                d.percent_complete_delta,
                d.total_slack_delta,
                d.baseline_start_delta_days,
                d.baseline_finish_delta_days,
                "Y" if d.became_critical else "",
                "Y" if d.dropped_off_critical else "",
            ]
        )
    _autosize_columns(ws)


def _write_dcma(ws: Worksheet, dcma) -> None:
    ws.append(["#", "Metric", "Value", "Unit", "Threshold", "Status"])
    _style_header_row(ws)

    for m in dcma.metrics:
        row_idx = ws.max_row + 1
        ws.append(
            [
                m.number,
                m.name,
                m.value,
                m.unit,
                f"{m.comparison}{m.threshold}",
                "PASS" if m.passed else "FAIL",
            ]
        )
        status_cell = ws.cell(row=row_idx, column=6)
        if m.passed:
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        else:
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT

    _autosize_columns(ws)


def _write_manipulation(ws: Worksheet, manip) -> None:
    ws.append(
        [
            "Confidence",
            "Category",
            "Pattern",
            "Severity",
            "Task UID",
            "Task Name",
            "Description",
        ]
    )
    _style_header_row(ws)

    conf_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    ordered = sorted(
        manip.findings,
        key=lambda f: (conf_rank.get(f.confidence, 9), -f.severity_score),
    )
    for f in ordered:
        row_idx = ws.max_row + 1
        ws.append(
            [
                f.confidence,
                f.category,
                f.pattern,
                f.severity_score,
                f.task_uid,
                f.task_name,
                f.description,
            ]
        )
        conf_cell = ws.cell(row=row_idx, column=1)
        if f.confidence == "HIGH":
            conf_cell.fill = HIGH_FILL
        elif f.confidence == "MEDIUM":
            conf_cell.fill = MEDIUM_FILL
        else:
            conf_cell.fill = LOW_FILL
        conf_cell.font = Font(bold=True)

    _autosize_columns(ws)


def _write_trend(ws: Worksheet, trend) -> None:
    ws.append(
        [
            "Update",
            "Status Date",
            "Project Finish",
            "Tasks",
            "Complete",
            "In Progress",
            "Not Started",
            "Critical Path Len",
            "Avg Float",
            "Min Float",
            "SPI",
            "BEI",
            "Manipulation Score",
            "Added Since Prior",
            "Removed Since Prior",
            "Completed Since Prior",
            "Finish Slip Since Prior (d)",
        ]
    )
    _style_header_row(ws)

    for dp in trend.data_points:
        ws.append(
            [
                dp.update_label,
                _fmt_date(dp.status_date),
                _fmt_date(dp.project_finish),
                dp.task_count,
                dp.tasks_complete,
                dp.tasks_in_progress,
                dp.tasks_not_started,
                dp.critical_path_task_count,
                dp.total_float_avg,
                dp.total_float_min,
                dp.spi,
                dp.bei,
                dp.manipulation_score,
                dp.tasks_added_since_prior,
                dp.tasks_removed_since_prior,
                dp.tasks_completed_since_prior,
                dp.finish_slip_since_prior_days,
            ]
        )
    _autosize_columns(ws)
